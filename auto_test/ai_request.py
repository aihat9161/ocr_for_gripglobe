import requests
import json
import os
from datetime import datetime
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import base64

# OpenAIのAPIキーを設定 (環境変数から読み込み)
OPENAI_API_KEY = os.getenv("OPENAI_APIKEY")

# 生成されたJSONファイルの保存先フォルダを設定 (絶対パス)
json_folder = os.path.join(os.path.dirname(__file__), "jsons")

# jsonsフォルダが存在しない場合は作成
if not os.path.exists(json_folder):
    os.makedirs(json_folder)

# OpenAI APIを呼び出す関数
def call_openai_api(base64_content):
    try:
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        }
        data = {
            "model": "gpt-4o-2024-08-06",
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": "この画像から、取引合計金額、取引日付（年、月、日を一つのフィールドにまとめる）、取引相手を抽出し、JSON形式で返してください。以下の形式で応答してください：\n{\"amount\": 取引合計金額, \"date\": \"取引日付\", \"trading_partner\": \"取引相手\"}"
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_content}"
                            }
                        }
                    ]
                }
            ],
            "max_tokens": 2000
        }
        response = requests.post("https://api.openai.com/v1/chat/completions", headers=headers, json=data)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"OpenAI API呼び出しエラー: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"OpenAI API呼び出し中にエラーが発生しました: {e}")
        return None

# APIレスポンスから情報を抽出する関数
def extract_info_from_response(response_data):
    if response_data is None or 'choices' not in response_data:
        return None
    
    content = response_data['choices'][0]['message']['content']
    
    try:
        # JSONデータを探す
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0)
            extracted_info = json.loads(json_str)
            return extracted_info
        else:
            print("JSONデータが見つかりませんでした。")
            return None
    except json.JSONDecodeError:
        print("JSONの解析に失敗しました。")
        return None

# 画像ファイルの形式をチェックする関数
def check_image_format(base64_content):
    try:
        image_data = base64.b64decode(base64_content)
        # PNGの場合
        if image_data.startswith(b'\x89PNG\r\n\x1a\n'):
            return True
        # JPEGの場合
        elif image_data.startswith(b'\xff\xd8'):
            return True
        # GIFの場合
        elif image_data.startswith(b'GIF87a') or image_data.startswith(b'GIF89a'):
            return True
        # WEBPの場合
        elif image_data.startswith(b'RIFF') and image_data[8:12] == b'WEBP':
            return True
        else:
            print("サポートされていない画像形式です。")
            return False
    except:
        print("画像データの検証中にエラーが発生しました。")
        return False

# JSON形式で出力して保存する関数
def save_response_as_json(response_data):
    if response_data is None:
        return None
    
    extracted_info = extract_info_from_response(response_data)
    if extracted_info is None:
        print("情報の抽出に失敗しました。")
        return None
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_file_name = f"response_{timestamp}.json"
    json_file_path = os.path.join(json_folder, json_file_name)
    
    with open(json_file_path, "w", encoding='utf-8') as json_file:
        json.dump(extracted_info, json_file, ensure_ascii=False, indent=4)
    
    print(f"レスポンスが {json_file_path} に保存されました。")
    return extracted_info

# データの完全性をチェックする関数
def check_data_completeness(data):
    return all(data.get(field) for field in ['amount', 'date', 'trading_partner'])

# Excelファイルに書き込む関数
def write_to_excel(data_list):
    excel_file = "json_check.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "json_check"

    # ヘッダーを設定
    headers = ["Filename", "JSON All", "JSON Three Elements", "Amount", "Date", "Trading Partner", "Completeness"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # データを書き込む
    for row, (filename, data) in enumerate(data_list, start=2):
        ws.cell(row=row, column=1, value=filename)
        ws.cell(row=row, column=2, value=json.dumps(data, ensure_ascii=False))
        ws.cell(row=row, column=3, value=f"{data.get('amount', '')}, {data.get('date', '')}, {data.get('trading_partner', '')}")
        ws.cell(row=row, column=4, value=data.get('amount', ''))
        ws.cell(row=row, column=5, value=data.get('date', ''))
        ws.cell(row=row, column=6, value=data.get('trading_partner', ''))
        
        # データの完全性をチェックし、結果を書き込む
        is_complete = check_data_completeness(data)
        completeness_cell = ws.cell(row=row, column=7, value='True' if is_complete else 'False')
        
        # セルの背景色を設定
        if is_complete:
            completeness_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        else:
            completeness_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 列幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # 最大幅を50に制限
        ws.column_dimensions[column_letter].width = adjusted_width

    # ファイルを保存
    wb.save(excel_file)
    print(f"データが {excel_file} に書き込まれました。")
    
# Base64エンコードされたコンテンツのリストを受け取り、処理する関数
def process_base64_list(base64_list):
    all_data = []
    for filename, base64_content in base64_list:
        if check_image_format(base64_content):
            response = call_openai_api(base64_content)
            extracted_info = save_response_as_json(response)
            if extracted_info:
                all_data.append((filename, extracted_info))
        else:
            print(f"ファイル {filename} はサポートされていない形式です。スキップします。")
    
    if all_data:
        write_to_excel(all_data)
    else:
        print("処理可能なデータがありませんでした。")